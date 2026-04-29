"""
报告生成模块

导出核对结果为 Excel 文件，包含：
- 汇总对照表（差异标红）
- 各 PDF 详情 Sheet
- 未找到项
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from backend.models import IndicatorResult, Indicator
from backend.config import OUTPUT_DIR

logger = logging.getLogger(__name__)

# 尝试导入 win32com
try:
    import win32com.client
    from win32com.client import constants
    HAS_WIN32COM = True
except ImportError:
    HAS_WIN32COM = False
    logger.warning("win32com 未安装，线程化批注功能不可用。请执行: pip install pywin32")

# 样式定义
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
DIFF_FILL = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
DIFF_FONT = Font(color="CC0000", bold=True)
MATCH_FILL = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
NOT_FOUND_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

STATUS_MAP = {
    "已核对": ("🟢", PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")),
    "未核对": ("🟡", PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")),
}


def generate_report(
    results: list[IndicatorResult],
    pdf_names: list[str],
    output_path: str | Path | None = None,
) -> Path:
    """
    生成核对结果 Excel 报告。

    Args:
        results: 比对结果列表
        pdf_names: PDF 文件名列表
        output_path: 输出路径，默认 output/报告数据核对结果.xlsx

    Returns:
        生成的文件路径
    """
    if output_path is None:
        output_path = OUTPUT_DIR / "报告数据核对结果.xlsx"
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # Sheet 1: 汇总对照表
    _create_summary_sheet(wb, results, pdf_names)

    # Sheet 2+: 各 PDF 详情
    for pdf_name in pdf_names:
        _create_pdf_detail_sheet(wb, results, pdf_name)

    # 最后一个 Sheet: 未找到项
    _create_not_found_sheet(wb, results, pdf_names)

    # 删除默认的空 Sheet
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    wb.save(str(output_path))
    logger.info(f"报告已生成: {output_path}")
    return output_path


def _add_threaded_comment_win32com(
    excel_path: str | Path,
    output_path: str | Path,
    row_idx: int,
    col_idx: int,
    note_text: str,
) -> None:
    """
    使用 win32com 在指定单元格添加线程化批注（现代批注）。

    优先使用 AddCommentThreaded（线程化批注），若不支持则降级为 AddComment。

    Args:
        excel_path: 源 Excel 文件路径
        output_path: 输出 Excel 文件路径
        row_idx: 单元格行号（1-based）
        col_idx: 单元格列号（1-based）
        note_text: 批注文本内容
    """
    if not HAS_WIN32COM:
        logger.warning("win32com 不可用，跳过线程化批注写入")
        return

    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(excel_path))
        ws = wb.ActiveSheet

        # 清除该单元格可能存在的旧批注
        ws.Range(
            ws.Cells(row_idx, col_idx),
            ws.Cells(row_idx, col_idx),
        ).ClearComments()

        target_range = ws.Cells(row_idx, col_idx)

        try:
            # 优先尝试线程化批注（现代批注）
            target_range.AddCommentThreaded(note_text)
            logger.debug(f"已添加线程化批注: 行{row_idx} 列{col_idx}")
        except AttributeError:
            # 降级：使用传统批注
            logger.warning("当前 Excel 版本不支持线程化批注，降级为传统批注")
            try:
                target_range.AddComment(note_text)
                logger.debug(f"已添加传统批注: 行{row_idx} 列{col_idx}")
            except Exception as e:
                logger.warning(f"添加传统批注失败 (行{row_idx} 列{col_idx}): {e}")
        except Exception as e:
            logger.warning(f"添加线程化批注失败 (行{row_idx} 列{col_idx}): {e}")
            # 尝试降级
            try:
                target_range.AddComment(note_text)
                logger.debug(f"降级后已添加传统批注: 行{row_idx} 列{col_idx}")
            except Exception as e2:
                logger.warning(f"降级添加传统批注也失败 (行{row_idx} 列{col_idx}): {e2}")

        # 保存为 xlsx 格式 (FileFormat=51)
        wb.SaveAs(str(output_path), FileFormat=51)

    except Exception as e:
        logger.warning(f"win32com 批注处理失败: {e}")
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def generate_colored_original_excel(
    results: list[IndicatorResult],
    excel_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """
    在原表基础上生成标色版本。
    绿色（已核对）
    黄色（未核对）

    批注使用 win32com 实现线程化批注（现代批注），
    若 Excel 版本不支持则自动降级为传统批注。
    """
    if output_path is None:
        output_path = OUTPUT_DIR / "标色原表.xlsx"
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # ---- 第一步：使用 openpyxl 处理标色逻辑 ----
    wb = load_workbook(str(excel_path))
    ws = wb.active

    # 尝试映射列名到列索引
    col_name_to_idx = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_val = str(ws.cell(row=1, column=col_idx).value or "").strip()
        if cell_val:
            col_name_to_idx[cell_val] = col_idx

    font_colors = {
        "已核对": Font(color="000000"),     # 不改变字体
        "未核对": Font(color="FF6600"),    # 橙色字体（不改背景色，避免覆盖原有分类颜色）
    }

    # 收集需要添加批注的信息（行号、列号、批注内容）
    comment_targets: list[tuple[int, int, str]] = []

    for result in results:
        ind = result.indicator
        if not ind.row_index:
            continue

        row_idx = ind.row_index
        col_idx = col_name_to_idx.get(ind.col_name) if ind.col_name else None

        # 如果需要根据匹配情况推断"未核对"的状态，可以选择推断
        # 但为保持与 UI 状态同步，这里直接以 review_status 决定颜色
        # 由于前端名为"未核对"，这里与它对应
        status_key = ind.review_status
        # 防止意外的非法状态，默认当做未核对
        if status_key not in font_colors:
            status_key = "未核对"

        font_color = font_colors[status_key]

        if col_idx:
            target_cell = ws.cell(row=row_idx, column=col_idx)
            target_cell.font = font_color
        else:
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=c).font = font_color
            target_cell = ws.cell(row=row_idx, column=1)

        # 收集批注信息（使用 openpyxl 的列索引）
        if ind.note and ind.note.strip():
            # 确定批注目标列：优先使用 col_idx，否则使用第1列
            comment_col = col_idx if col_idx else 1
            comment_targets.append((row_idx, comment_col, ind.note.strip()))

    # 先保存标色后的中间文件（使用临时路径）
    temp_path = output_path.with_suffix(".tmp.xlsx")
    wb.save(str(temp_path))
    wb.close()

    # ---- 第二步：使用 win32com 添加线程化批注 ----
    if comment_targets and HAS_WIN32COM:
        _add_threaded_comments_batch(temp_path, output_path, comment_targets)
        # 删除临时文件
        try:
            temp_path.unlink()
        except Exception:
            pass
    else:
        # 没有批注或 win32com 不可用，直接重命名临时文件为最终输出
        if not HAS_WIN32COM and comment_targets:
            logger.warning("win32com 不可用，批注将以 openpyxl 旧版批注形式保留")
            # 重新加载并添加旧版批注
            wb2 = load_workbook(str(temp_path))
            ws2 = wb2.active
            for row_idx, col_idx, note_text in comment_targets:
                comment = Comment(text=note_text, author="核对人员")
                comment.width = 300
                comment.height = 100
                ws2.cell(row=row_idx, column=col_idx).comment = comment
            wb2.save(str(output_path))
            wb2.close()
            try:
                temp_path.unlink()
            except Exception:
                pass
        else:
            # 没有批注需要添加，直接重命名
            import shutil
            shutil.move(str(temp_path), str(output_path))

    logger.info(f"标色原表已生成: {output_path}")
    return output_path


def _add_threaded_comments_batch(
    temp_path: Path,
    output_path: Path,
    comment_targets: list[tuple[int, int, str]],
) -> None:
    """
    批量使用 win32com 添加线程化批注。

    Args:
        temp_path: openpyxl 处理后的临时文件路径
        output_path: 最终输出文件路径
        comment_targets: 列表，每项为 (row_idx, col_idx, note_text)
    """
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(temp_path))
        ws = wb.ActiveSheet

        for row_idx, col_idx, note_text in comment_targets:
            try:
                target_range = ws.Cells(row_idx, col_idx)

                # 清除可能存在的旧批注
                target_range.ClearComments()

                try:
                    # 优先尝试线程化批注（现代批注）
                    target_range.AddCommentThreaded(note_text)
                    logger.debug(f"已添加线程化批注: 行{row_idx} 列{col_idx}")
                except AttributeError:
                    # 降级：使用传统批注
                    logger.warning("当前 Excel 版本不支持线程化批注，降级为传统批注")
                    try:
                        target_range.AddComment(note_text)
                        logger.debug(f"已添加传统批注: 行{row_idx} 列{col_idx}")
                    except Exception as e:
                        logger.warning(f"添加传统批注失败 (行{row_idx} 列{col_idx}): {e}")
                except Exception as e:
                    logger.warning(f"添加线程化批注失败 (行{row_idx} 列{col_idx}): {e}")
                    # 尝试降级
                    try:
                        target_range.AddComment(note_text)
                        logger.debug(f"降级后已添加传统批注: 行{row_idx} 列{col_idx}")
                    except Exception as e2:
                        logger.warning(f"降级添加传统批注也失败 (行{row_idx} 列{col_idx}): {e2}")

            except Exception as e:
                logger.warning(f"批注处理异常 (行{row_idx} 列{col_idx}): {e}")
                continue

        # 保存为 xlsx 格式 (FileFormat=51)
        wb.SaveAs(str(output_path), FileFormat=51)
        logger.info(f"win32com 批注处理完成，已保存至: {output_path}")

    except Exception as e:
        logger.warning(f"win32com 批注处理失败: {e}")
        # 如果失败，尝试将临时文件复制为输出文件
        import shutil
        shutil.copy2(str(temp_path), str(output_path))
        logger.info(f"win32com 处理失败，已使用 openpyxl 版本: {output_path}")
    finally:
        if wb is not None:
            try:
                wb.Close(SaveChanges=False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _apply_header_style(ws, row: int, max_col: int):
    """为表头行应用样式。"""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """自动调整列宽。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val.encode("gbk", errors="replace")))
            except Exception:
                max_len = max(max_len, len(str(cell.value or "")))
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def _create_summary_sheet(wb: Workbook, results: list[IndicatorResult], pdf_names: list[str]):
    """创建汇总对照表 Sheet。"""
    ws = wb.active
    ws.title = "汇总对照表"

    # 表头：年份、核对情况、一二三级标题、数据
    headers = [
        "序号", "年份", "一级标题", "二级标题", "三级标题（指标名称）",
        "Excel目标值", "单位", "核对情况", "匹配值", "是否一致",
        "匹配PDF", "匹配页码", "数据来源", "批注",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    # 数据行
    for row_idx, result in enumerate(results, 2):
        ind = result.indicator

        # 序号
        ws.cell(row=row_idx, column=1, value=ind.id + 1)

        # 年份
        ws.cell(row=row_idx, column=2, value=ind.year or "")

        # 一级标题
        ws.cell(row=row_idx, column=3, value=ind.category1 or "")

        # 二级标题
        ws.cell(row=row_idx, column=4, value=ind.category2 or "")

        # 三级标题（指标名称）
        ws.cell(row=row_idx, column=5, value=ind.name)

        # Excel目标值（非数字则显示原始文本）
        excel_value = ind.raw_target if (not ind.is_numeric and ind.raw_target) else ind.target_value
        ws.cell(row=row_idx, column=6, value=excel_value)

        # 单位
        ws.cell(row=row_idx, column=7, value=ind.unit or "")

        # 核对情况
        status_icon, status_fill = STATUS_MAP.get(ind.review_status, ("🔵", None))
        status_cell = ws.cell(row=row_idx, column=8, value=f"{status_icon} {ind.review_status}")
        if status_fill:
            status_cell.fill = status_fill

        # 从 best_matches 中提取最佳匹配信息
        best_match = None
        best_confidence = -1
        for pdf_name, match in result.best_matches.items():
            if match and match.confidence > best_confidence:
                best_match = match
                best_confidence = match.confidence

        # 匹配值
        if best_match and best_match.matched_value is not None:
            match_cell = ws.cell(row=row_idx, column=9, value=best_match.matched_value)
        else:
            match_cell = ws.cell(row=row_idx, column=9, value="未找到")

        # 是否一致
        if best_match:
            is_match = best_match.is_match
            consistent_cell = ws.cell(row=row_idx, column=10, value="✓ 一致" if is_match else "✗ 不一致")
            if is_match:
                consistent_cell.fill = MATCH_FILL
            else:
                consistent_cell.fill = DIFF_FILL
                consistent_cell.font = DIFF_FONT
        else:
            consistent_cell = ws.cell(row=row_idx, column=10, value="未匹配")
            consistent_cell.fill = NOT_FOUND_FILL

        # 匹配PDF
        ws.cell(row=row_idx, column=11, value=ind.matched_pdf_name or (best_match.pdf_name if best_match else ""))

        # 匹配页码
        ws.cell(row=row_idx, column=12, value=ind.matched_page or (best_match.page_number if best_match else ""))

        # 数据来源（source_file）
        ws.cell(row=row_idx, column=13, value=ind.source_file or "")

        # 批注
        note_cell = ws.cell(row=row_idx, column=14, value=ind.note or "")
        if ind.note and ind.note.strip():
            # 在指标名称单元格也添加 Excel 批注（作为悬停提示）
            name_cell = ws.cell(row=row_idx, column=5)
            comment = Comment(
                text=ind.note.strip(),
                author="核对人员",
            )
            comment.width = 300
            comment.height = 100
            name_cell.comment = comment

        # 行边框
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    _auto_width(ws)


def _create_pdf_detail_sheet(wb: Workbook, results: list[IndicatorResult], pdf_name: str):
    """创建单个 PDF 详情 Sheet。"""
    short_name = Path(pdf_name).stem[:28]
    ws = wb.create_sheet(title=short_name)

    headers = ["指标名称", "Excel值", "匹配值", "原始字符串", "页码",
               "置信度", "是否一致", "差值", "上下文原文", "批注"]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        matches = result.matches.get(pdf_name, [])
        ind = result.indicator
        if not matches:
            ws.cell(row=row_idx, column=1, value=ind.name)
            ws.cell(row=row_idx, column=2, value=ind.target_value)
            ws.cell(row=row_idx, column=3, value="未找到")
            ws.cell(row=row_idx, column=3).fill = NOT_FOUND_FILL
            ws.cell(row=row_idx, column=10, value=ind.note or "")
            row_idx += 1
            continue

        for m in matches:
            ws.cell(row=row_idx, column=1, value=ind.name)
            ws.cell(row=row_idx, column=2, value=ind.target_value)
            ws.cell(row=row_idx, column=3, value=m.matched_value)
            ws.cell(row=row_idx, column=4, value=m.matched_value_raw)
            ws.cell(row=row_idx, column=5, value=m.page_number)
            ws.cell(row=row_idx, column=6, value=f"{m.confidence}%")

            match_cell = ws.cell(row=row_idx, column=7, value="✓ 一致" if m.is_match else "✗ 不一致")
            if not m.is_match:
                match_cell.fill = DIFF_FILL
                match_cell.font = DIFF_FONT
            else:
                match_cell.fill = MATCH_FILL

            ws.cell(row=row_idx, column=8, value=m.difference)
            ws.cell(row=row_idx, column=9, value=m.context)
            ws.cell(row=row_idx, column=10, value=ind.note or "")

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER

            row_idx += 1

    _auto_width(ws)


def _create_not_found_sheet(wb: Workbook, results: list[IndicatorResult], pdf_names: list[str]):
    """创建未找到项 Sheet。"""
    ws = wb.create_sheet(title="未找到项")

    headers = ["指标名称", "Excel值", "单位", "未匹配的PDF", "批注"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _apply_header_style(ws, 1, len(headers))

    row_idx = 2
    for result in results:
        ind = result.indicator
        missing_pdfs = []
        for pdf_name in pdf_names:
            best = result.best_matches.get(pdf_name)
            if best is None:
                missing_pdfs.append(pdf_name)

        if missing_pdfs:
            ws.cell(row=row_idx, column=1, value=ind.name)
            excel_value = ind.raw_target if (not ind.is_numeric and ind.raw_target) else ind.target_value
            ws.cell(row=row_idx, column=2, value=excel_value)
            ws.cell(row=row_idx, column=3, value=ind.unit or "")
            ws.cell(row=row_idx, column=4, value=", ".join(missing_pdfs))
            ws.cell(row=row_idx, column=5, value=ind.note or "")

            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).border = THIN_BORDER
                ws.cell(row=row_idx, column=col).fill = NOT_FOUND_FILL

            row_idx += 1

    _auto_width(ws)
